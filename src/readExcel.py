# -*- coding: utf-8 -*-

import os
from openpyxl import Workbook
# from openpyxl.utils import get_column_letter

fileName = 'test.xlsx'
outputFile = os.path.join('out', fileName)
if os.path.exists(outputFile):
    os.remove(outputFile)

wb = Workbook()
wsOfDefault = wb.active
wsOfDefault.title = "测试"

# wsOfTest = wb.create_sheet('test', 1)
# wsOfTest.title = 'test'

for row in range(1, 10):
    wsOfDefault.append(range(10))

# ws2 = wb.create_sheet(title="Pi")
# ws2['F5'] = 3.14

# ws3 = wb.create_sheet(title="字母")
# for row in range(10, 20):
#     for col in range(17, 24):
#         # _ = ws3.cell(column=col, row=row, value="%s" % get_column_letter(col))
#         ws3.cell(column=col, row=row, value="%s" % get_column_letter(col))

wb.save(filename=outputFile)
