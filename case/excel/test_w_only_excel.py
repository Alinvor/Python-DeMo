# -*- coding: utf-8 -*-

# import os
from openpyxl import Workbook
from openpyxl.cell import WriteOnlyCell
from openpyxl.comments import Comment
from openpyxl.styles import Font
from openpyxl.utils import units
from com.dvsnier.dir.common_dir import generate_complex_or_fmt_file_name
from openpyxl.worksheet.properties import PageSetupProperties


def write_to_excel(file_name):
    'the write data to excel file'
    if file_name is None or len(file_name.strip()) == 0:
        raise ValueError('the current file name is invaild.')

    wb = Workbook(write_only=True)  # 只写模式，工作簿
    work_sheet_of_default = wb.create_sheet()
    work_sheet_of_default.title = u"只写测试0"  # 当前工作表title
    # 只写模式，Cell
    write_only_cell = WriteOnlyCell(work_sheet_of_default, value="hello world")
    write_only_cell.font = Font(name='Courier', size=20)
    # 注释具有text属性和author属性，必须同时设置
    write_only_cell.comment = Comment(text="cell comment",
                                      author="Author's Name")
    write_only_cell.comment.width = units.points_to_pixels(250)
    write_only_cell.comment.height = 250
    # 在只写工作簿中, 只能使用添加行 append(), 无法使用 cell()或在任意位置写入(或读取)单元 iter_rows()
    work_sheet_of_default.append([None, 3.14, write_only_cell])
    # 工作表注释
    work_sheet_of_comment = Comment(text="ws comment", author=u"资深好男人")
    work_sheet_of_default.comment = work_sheet_of_comment
    #
    # 工作表可用属性
    #
    # “enableFormatConditionsCalculation”
    # “filterMode”
    # “published”
    # “syncHorizontal”
    # “syncRef”
    # “syncVertical”
    # “transitionEvaluation”
    # “transitionEntry”
    # “tabColor”
    wsprops = work_sheet_of_default.sheet_properties  # 工作表属性
    wsprops.tabColor = "1072BA"
    wsprops.filterMode = False
    wsprops.pageSetUpPr = PageSetupProperties(fitToPage=True,
                                              autoPageBreaks=False)
    wsprops.outlinePr.summaryBelow = False
    wsprops.outlinePr.applyStyles = True
    wsprops.pageSetUpPr.autoPageBreaks = True

    wb.save(filename=file_name)


if __name__ == "__main__":
    u'''只写Excel 文件'''
    file_name = generate_complex_or_fmt_file_name('excel', 'excel.xlsx')
    #
    # 写xlsx 文件
    #
    write_to_excel(file_name)
