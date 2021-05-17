# -*- coding: utf-8 -*-

from com.dvsnier.dir.common_dir import generate_complex_or_fmt_file_name
from copy import copy
import datetime
# import os
from openpyxl import Workbook
# from openpyxl.styles import Alignment, Border, Font, PatternFill, Protection, Side, colors
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side, colors
from openpyxl.styles.fills import GradientFill
from openpyxl.styles.named_styles import NamedStyle
from openpyxl.reader.excel import load_workbook


def ready_to_excel(file_name):
    'the read data from excel file'
    if file_name is None or len(file_name.strip()) == 0:
        raise ValueError('the current file name is invaild.')
    wb = load_workbook(file_name)
    sheet_names = wb.sheetnames
    # names = []
    for name in sheet_names:
        print(name.encode('utf-8'))
        # names.append(name.encode('utf-8'))
    # print(names)


def write_to_excel(file_name):
    'the write data to excel file'
    if file_name is None or len(file_name.strip()) == 0:
        raise ValueError('the current file name is invaild.')

    wb = Workbook()  # 工作簿
    work_sheet_of_default = wb.active  # 当前工作表,等价于wb.active(0)
    work_sheet_of_default.title = u"测试0"  # 当前工作表title
    # 单元格赋值
    work_sheet_of_default['A8'] = 'A8'
    ws_cell = work_sheet_of_default.cell(row=9, column=1, value='A9')
    ws_cell.value = 'A99'
    # 合并/撤销单元格
    work_sheet_of_default.merge_cells('B2:C2')
    # work_sheet_of_default.merge_cells(start_row=2,
    #                                   start_column=2,
    #                                   end_row=2,
    #                                   end_column=3)
    work_sheet_of_default['B2'] = u'合并的单元格'
    # work_sheet_of_default.unmerge_cells('B2:C2')
    # 区间访问
    for row in range(1, 10):
        work_sheet_of_default.append(range(10))

    # 创建新的工作表 WorkSheet
    work_sheet_of_test = wb.create_sheet()
    # work_sheet_of_test = wb.create_sheet(title='test',
    #                                      index=1)  # title 和index 可不写省略
    work_sheet_of_test.title = u'test1'
    work_sheet_of_test.sheet_properties.tabColor = "1072BA"  # 设置选项卡背景色
    # 单元格取值
    # for row in work_sheet_of_test.values():
    #     for value in row:
    #         print(value)
    # 行列访问
    row_3 = work_sheet_of_test['3']
    row_range = work_sheet_of_test['3:5']
    row_str = 'row: %s\t\t%s' % (row_3, row_range)
    row_str = 'row pass'
    print(row_str)
    column_c = work_sheet_of_test['C']
    column_ragne = work_sheet_of_test['C:D']
    column_str = 'column: %s\t\t%s' % (column_c, column_ragne)
    column_str = 'column pass'
    print(column_str)
    # 迭代访问，迭代行
    for row in work_sheet_of_test.iter_rows(min_row=1, max_row=2, max_col=3):
        for cell in row:
            # print(cell)
            pass
    # 迭代访问，迭代列
    for col in work_sheet_of_test.iter_cols(min_row=1, max_row=2,
                                            max_col=3):  # 只读模式不可用
        for cell in col:
            # print (cell)
            pass

    work_sheet_pi = wb.create_sheet(title=u"Pi")
    work_sheet_pi['F5'] = 3.14
    # 日期格式
    work_sheet_pi['A3'] = datetime.datetime.now()
    print(work_sheet_pi['A3'].number_format)
    # 公式
    work_sheet_pi['B2'] = "=SUM(1, 1)"
    # 折叠行/列
    work_sheet_pi.row_dimensions.group(10, 12, hidden=True)
    work_sheet_pi.column_dimensions.group('F', 'G', hidden=True)

    work_sheet_ascii = wb.create_sheet(title=u"字母")
    for row in range(10, 20):
        for col in range(10, 20):
            work_sheet_ascii.cell(column=col,
                                  row=row,
                                  value=str("%s%s" % (row, col)))
    # minimum_bounding_range = work_sheet_ascii.calculate_dimension()
    # print(minimum_bounding_range)
    work_sheet_style = wb.create_sheet(title=u"style")
    #
    # 样式默认值
    #
    font = Font(name='Calibri',
                size=11,
                bold=False,
                italic=False,
                vertAlign=None,
                underline='none',
                strike=False,
                color='FF000000')
    fill = PatternFill(fill_type=None,
                       start_color='FFFFFFFF',
                       end_color='FF000000')
    # border = Border(left=Side(border_style=None, color='FF000000'),
    #                 right=Side(border_style=None, color='FF000000'),
    #                 top=Side(border_style=None, color='FF000000'),
    #                 bottom=Side(border_style=None, color='FF000000'),
    #                 diagonal=Side(border_style=None, color='FF000000'),
    #                 diagonal_direction=0,
    #                 outline=Side(border_style=None, color='FF000000'),
    #                 vertical=Side(border_style=None, color='FF000000'),
    #                 horizontal=Side(border_style=None, color='FF000000'))
    # alignment = Alignment(horizontal='general',
    #                       vertical='bottom',
    #                       text_rotation=0,
    #                       wrap_text=False,
    #                       shrink_to_fit=False,
    #                       indent=0)
    # number_format = 'General'
    # protection = Protection(locked=True, hidden=False)
    #
    # 单元格样式在对象之间共享，并且一旦分配它们便无法更改
    #
    # https://openpyxl.readthedocs.io/en/2.6/styles.html#cell-styles-and-named-styles
    ws_b5 = work_sheet_style['B5']
    font1 = copy(font)  # 样式复制
    font1.name = 'Arial'
    ws_b5.font = font1
    ws_b5.value = 'ws_b5'

    ws_d6 = work_sheet_style['D6']
    font2 = Font(name='Tahoma', color=colors.RED, italic=True, size=16)
    ws_d6.font = font2
    ws_d6.value = 'ws_d6'
    #
    # 样式合并单元格
    #
    work_sheet_style.merge_cells('B8:F10')
    top_left_cell = work_sheet_style['B8']
    top_left_cell.value = u"样式合并单元格"
    # 边
    thin = Side(border_style="thin", color="000000")
    double = Side(border_style="double", color="ff0000")
    # 边框
    top_left_cell.border = Border(top=double,
                                  left=thin,
                                  right=thin,
                                  bottom=double)
    # 填充
    top_left_cell.fill = PatternFill("solid", fgColor="DDDDDD")
    fill = GradientFill(stop=("000000", "FFFFFF"))
    top_left_cell.fill = fill
    top_left_cell.font = Font(b=True, color="FF0000")
    # 对齐模式
    top_left_cell.alignment = Alignment(horizontal="center", vertical="center")

    #
    # 页面设置
    #
    work_sheet_style.page_setup.orientation = work_sheet_style.ORIENTATION_LANDSCAPE
    work_sheet_style.page_setup.paperSize = work_sheet_style.PAPERSIZE_TABLOID
    work_sheet_style.page_setup.fitToWidth = 1
    work_sheet_style.page_setup.fitToHeight = 0

    #
    # 与单元样式相反，命名样式是可变的
    #
    # 当您想一次将格式应用于许多不同的单元格时，它们很有意义。注意 将命名样式分配给单元后，对该样式的其他更改将 不会影响该单元。
    # 一旦将命名样式注册到工作簿中，就可以简单地通过名称来引用它。
    #

    # 创建样式,注册与指定
    high_light_style = NamedStyle(name="highlight")
    high_light_style.font = Font(bold=True, size=20, color=colors.RED)
    boder_style = Side(style='thick', color="000000")
    high_light_style.border = Border(left=boder_style,
                                     top=boder_style,
                                     right=boder_style,
                                     bottom=boder_style)
    wb.add_named_style(high_light_style)  # 工作簿注册样式，命名样式在首次分配给单元时也会自动注册
    ws_g8 = work_sheet_style['G8']
    ws_g8.style = high_light_style
    ws_g8.value = 'ws_g8'
    ws_h6 = work_sheet_style['H6']
    ws_h6.style = 'highlight'  # 工作簿注册样式后，仅使用名称指定样式
    ws_h6.value = 'ws_h6'
    #
    # 使用内置样式
    # 请参考:
    #       https://openpyxl.readthedocs.io/en/2.6/styles.html#using-builtin-styles
    #

    wb.save(filename=file_name)


if __name__ == "__main__":
    u'''读写Excel 文件'''
    file_name = generate_complex_or_fmt_file_name('excel', 'excel.xlsx')
    #
    # 写xlsx 文件
    #
    write_to_excel(file_name)
    #
    # 读xlsx 文件
    #
    # ready_to_excel(file_name)
